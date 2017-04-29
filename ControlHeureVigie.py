
import pyexcel as pe #permet de manipuler les données de classeurs xls,xlsx
import openpyxl  #permet de lire et d'enregistrer des xlsx en gardant le format et les formules
from openpyxl.styles.borders import Border, Side #permet de gérer les bordures des cellules
import os #permet de changer de dossier qqsoit l'os
from inspect import getsourcefile
from datetime import date, timedelta #gère les dates et durées
import random #fait des nombres aléatoires
import csv #gère les fichiers csv



def trouve_dossier(nom): #renvoie le chemin complet du dossier "nom" même sur pc virtualisé
    chemin_script=os.path.abspath(getsourcefile(lambda:0)) #trouve le chemin complet du script python
    parent=os.path.dirname(chemin_script) #renvoie le dossier dans lequel se trouve le script
    dossier=parent
    trouve=""

    while not nom in parent: #cherche le dossier de façon récursive en parcourant tous les sous-dossier 
        liste=os.listdir(parent)
        for f in liste:
            if nom in f:
                trouve=os.path.join(parent,f) #si trouvé, recrée le chemin complet
                break
        dossier=parent #sinon remonte au dossier parent précédent
        parent=os.path.dirname(parent)
        if parent==dossier : #teste si on est à la racine, dans ce ca, renvoie un message d'erreur et arrête de chercher
            print("Le dossier où se trouve les plannings n'a pas été trouvé dans : ",dossier)
            break
        if trouve!="":
            break
    return trouve


def liste_plannings(annee):
    #renvoie une liste de nom des plannings en fonction de l'année
    l=list()
    with open("data.csv",encoding='utf-8') as data:
            readCSV = csv.reader(data, delimiter=',')
            for row in readCSV:
##                print(row)
                if "equipe" in row[0]:
                    dossier=row[1]
                    for equipe in ['A','B','C','D','E','F']:
                        chaine="Planning "+ equipe + " " + str(annee) + ".xlsm"
                        l.append(os.path.join(trouve_dossier(dossier),chaine))
                elif "detaches" in row[0]:
                    dossier=row[1]
                    if dossier!="":
                        chaine="planning CA "+str(annee)+".xls"
                        l.append(os.path.join(dossier,chaine))
                elif "stagiaires" in row[0]:
                    dossier=row[1]
                    if dossier!="":
                        chaine="planning stagiaires "+str(annee)+".xls"
                        l.append(os.path.join(dossier,chaine))
                elif row[0]=="fin": #indique que le fichier est terminé
                    break
    return l


        
def charge_planning(nom_planning):
    # charge les colonnes et lignes nécessaires de chaque planning pour n'avoir
    # que la partie des jours travaillés
    scol=3 #colonne début
    srow=1 #ligne début
    collim=18  #colonne fin
    if "planning CA" in nom_planning: #s'il s'agit du planning des détachés
        scol=1  #commence à la colonne 1
        srow=0  #à la ligne 0
        collim=25   #termine colonne 25 (au cas où il y en ait beaucoup)
        
    planning=pe.get_sheet(file_name=nom_planning,sheet_name="Planning",start_column=scol,column_limit=collim,start_row=srow,rowlimit=370)
    planning.name_columns_by_row(0) #utilise la première ligne pour faire référence aux colonnes
    #print(planning.row[0]) #pour test
    return planning

def extraire_vac(planning,trig,datedeb,datefin):
    #extrait les vac du plannig entre datedeb et datefin et renvoie une liste de listes [date,vacs]
    tab=list() #crée la liste tab
    lignedeb=(datedeb-date(datedeb.year,1,1)).days
    duree=datefin-datedeb
    nlignes=duree.days
    l=lignedeb
    if trig in planning.colnames:
        while l<=lignedeb+nlignes:                                      #################################################ajout <=
            if planning[l,trig]!="" :  #si la cellule n'est pas vide
                tab.append([planning[l,0],planning[l,trig]]) #ajoute à la liste le duo [date,vac]
            l+=1
        #print(tab)  #pour test
    return tab

def crée_liste_hdc(l_datvac,dic_forfaits,gamma,planstg):
    # renvoie une liste de
    # listes [date,heures double,heures standard,heures instructeur,heures simu,total,BA,BM]
    # en fonction des forfaits de dic_forfaits en ajoutant des heures en période été
    liste_hdc=list()
    for ligne in l_datvac:
        if ligne[1] in dic_forfaits.keys():
            nstg=planstg.nbr(ligne[0]) #détermine le nbr de stagiaires qui travaillent à la date liste[0]
            liste_hdc.append(rdm_forfait(ligne[0],dic_forfaits[ligne[1]],gamma,nstg))
    return liste_hdc
                                         

def rdm_forfait(dat,l_forfait,gamma,nstg):
    #renvoie une liste {date,heures dc,solo, inst,...} modifiée
    #en fonction d'un random et de la période été ou hiver
    l_rdm=list()
    l_rdm.append(dat) #premier élément de la liste est la date
    random.seed() #initialise le random
    delta=random.sample([-0.5,-0.25,0,0.25,0.5],1) #choisit un delta au hasard
    forfait=l_forfait[1] #affecte la valeur hiver (par défaut)
    nsem=dat.isocalendar()[1] #renvoie le numéro dela semaine
    if nsem<44 and nsem>13 : #si tds été, prend la valeur été
        forfait=l_forfait[0]
    if forfait=="" : #si la cellule est vide forfait prend la valeur 0
        forfait=0
    else :
        f=forfait
        forfait=float(f) #sinon forfait prend la valeur numérique

    if forfait>0 :
        forfait+=delta[0] #ajoute le delta
        forfait+=gamma # et le gamma (deduit les heures CDT)au forfait heure de ctle

    total=forfait
    tour=int(forfait*40/100)
    appbm=int(forfait*60/100)
    
    
    #et soustrait l'instruction
    if l_forfait[2]!="":
        if l_forfait[2]*nstg<forfait : #vérifie qu'il n'y a pas trop d'instruction
            l_forfait[2]*=nstg #multiplie le forfait instruction par le nbr de stagiaires
        forfait-=l_forfait[2]
        if l_forfait[2]==0:
            l_forfait[2]="" #si la valeur est nulle, n'affiche rien dans la case
    #et soustrait le simu
    if l_forfait[3]!="":
        forfait-=l_forfait[3]
    l_rdm.append(0) #ajoute 0 heure en double (ne gère pas les stagiaires)
    l_rdm.append(forfait) #ajoute le forfait (été ou hiver) modifié
    
    l_rdm.extend(l_forfait[2:]) #ajoute le reste des forfaits
    l_rdm.append(total) #ajoute le total des heures
    l_rdm.append(tour) #ajoute nombre heures tour
    l_rdm.append(appbm) #ajoute nombre heures ba/bm
    return l_rdm
    
def date_fin():
    #renvoie la date de fin du scanning des plannings (dernier jour du mois précédent)
    today=date.today()
    d=date(today.year,today.month,1)-timedelta(days=1)
    return d


#################################### classes ############################################################
class stagiaires:
    #charge le planning stagiaire et renvoie le nombre de stagiares présent le jour indiqué
    def __init__(self,nomfic=""):
        self.feuille=pe.get_sheet(file_name=nomfic,sheet_name="planning",start_column=3,start_row=1)

    def coldate(self,d):
        c=((d.month-1)*6) #détermine la colonne où se trouve la date
        return c
        
    def lignes(self,d): #renvoie les lignes début et fin de la date (corrige le pb lié aux cellules fusionnées)
        ldeb=-1
        col=self.coldate(d)
        lfin=self.feuille.number_of_rows()
        for l in range(1,lfin): #parcourt toutes les lignes
            valeur=self.feuille[l,col] #récupère la valeur de la cellule
            if valeur!="": #si la cellule date n'est pas vide
                if valeur.day==d.day : #si c'est la date recherchée, attribue le numéro de ligne à ldeb
                    ldeb=l
                elif ldeb!=-1 : #sinon, si ldeb a été trouvé, cherche jusqu'où aller jusqu'à la date suivante
                    lfin=l-1
                    break
        return [ldeb,lfin]
                

    def nbr(self,d): # renvoie le nombre de stagiaires présents le jour d
        coldat=self.coldate(d)
        [lignedeb,lignefin]=self.lignes(d)
        n=0
        for l in range(lignedeb,lignefin):
            for c in range(coldat+2,coldat+4):
                if self.feuille[l,c]!="" : # si un stagiaire dans une des colonnes st0, st1, st2 
                    n+=1 #rajoute le stagiaire
        return n

    
    
class nom_trig:
    #travaille avec le classeur noms et trigrammes et renvoie différentes listes
    def __init__(self,dossier=""):
        nomfic=os.path.join(dossier,"noms et trigrammes.xlsm")
        #charge les 6 premières colonnes de la feuille "noms et trigrammes"
        self.feuille=pe.get_sheet(file_name=nomfic,name_columns_by_row=0,start_column=0,column_limit=6)
              
    def liste_trig(self):
        #renvoie la liste des trigrammes concernés par les registres
        l=list()
        for r in self.feuille.rows():
            if r[5]==1 : #si le pc a accepté
                l.append(r[0])
        return l

    def nom_prenom(self,trig):
        #renvoie le nom prenom en fonction du trigramme
        np=""
        for r in self.feuille.rows():
            if r[0]==trig:
                np=r[4]
                break
        return np

    def nom(self,trig):
        #renvoie le nom raccourci en fonction du trig
        n=""
        for r in self.feuille.rows():
            if r[0]==trig :
                n=r[1]
                break
        return n

    def trig(self,nom):
        #renvoie le trigramme en fonction du nom raccourci
        t=""
        for r in self.feuille.rows():
            if r[1]==nom:
                t=r[0]
                break
        return t

    def gamma(self,trig):
        #renvoie un nbre d'h en plus ou en moins par exemple pour le CDT
        g=0
        for r in self.feuille.rows():
            if r[0]==trig and r[2]=="CDT":
                g=-2.5
                break
        return g

            
class dhc:
    #gère les décomptes heures de controle
    def __init__(self,trigramme,nom_prenom,dat=date.today()):
        self.datefin=dat
        self.trig=trigramme
        self.nom=nom_prenom
        self.dossier=self.dossier_dhc()
        self.nomfic=os.path.join(self.dossier,(self.trig+"_"+str(self.datefin.year)+(".xlsx")))
        self.wb=self.open_wb_dhc()
        self.datedeb=self.date_dhc() #date debut=date dernière mise à jour +1 jour
        
    def dossier_dhc(self):
        #vérifie que le dossier existe
        #si oui renvoie le dossier où se trouve les dhc sinon le crée et renvoie le dossier
        annee=self.datefin.year
        dossier=""
        with open("data.csv") as data:
            readCSV = csv.reader(data, delimiter=',')
            for row in readCSV:
                if row[0]=="Registres Controleurs":
                    dossier=row[1]
                    break
        if dossier=="":
            i=input("Impossible de trouver le dossier des registres de contrôleurs. Vérifier le fichier 'data.csv'.")
        else:
            drive=os.path.join(dossier,str(annee))
            #vérifie si le dossier existe, sinon le crée
            if not os.path.exists(drive):
                os.makedirs(drive)
            return drive
        
    def open_wb_dhc(self):
        #liste des mois pour remplir la case du DHC
        liste_mois=["","Janvier","Février","Mars","Avril","Mai","Juin","Juillet","Août","Septembre","Octobre","Novembre","Décembre"]
        #vérifie si le dhc existe sinon le crée
        if not os.path.exists(self.nomfic):
            trigwb = openpyxl.load_workbook('Registre_type.xlsx')
            trigwb.save(self.nomfic)
            ws0=trigwb["Data"] #selectionne la feuille "Données"
            dat=date(self.datefin.year,1,1)
            ws0['C1']=dat #affecte le 1er janvier de l'annee
            ws0['C3']=self.nom #affecte le nom_prenom du trig
            m=0
            for mois in liste_mois :
                if m!=0:
                    chaine=mois+" "+str(self.datefin.year)
                    trigwb[str(m)]["G1"]=chaine
                m+=1
            trigwb.save(self.nomfic) #sauve le classeur
        else:
            trigwb = openpyxl.load_workbook(self.nomfic)
        return trigwb

    def date_dhc(self):
        # renvoie la date de debut de mise à jour
        ws0=self.wb["Data"] #selectionne la feuille "Données"
        dat=ws0['C1'].value+timedelta(days=1)
        d=dat.date()
        return d

    def export_vers_dhc(self,l_vac):
        #exporte la liste l_vac vers le fichier dhc du controleur trig et sauvegarde

        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))
        
        medium_border = Border(left=Side(style='medium'),
                               right=Side(style='medium'),
                               top=Side(style='medium'),
                               bottom=Side(style='medium'))

        l_row_mois=[0,5,5,5,5,5,5,5,5,5,5,5,5] #crée une liste d'indice de lignes pour remplir chaque feuille mensuelle
        l_col=["","A","B","C","D","E","F","G"]
        for l in l_vac:
            col=1
            mois=l[0].month
            ws=self.wb[str(mois)] #sélectionne la feuille dont le nom est le mois considéré
            for val in l:
                v=ws.cell(row=l_row_mois[mois],column=col,value=val) #remplit les cellules avec les valeurs des vacs
                v.border=thin_border
                if col!=1 and col!=9: #sauf en colonne A et I
                    formule="=SUM("+l_col[col]+"5:"+l_col[col]+str(l_row_mois[mois])+")"
                    v=ws.cell(row=l_row_mois[mois]+1,column=col,value=formule) #en dessous de la ligne ajoute la formule pour calculer les totaux
                    v.border=medium_border
                elif col==1: #en colonne A
                    formule="TOTAL"
                    v=ws.cell(row=l_row_mois[mois]+1,column=col,value=formule)
                    v.border=medium_border
                col+=1
            l_row_mois[mois]+=1
        ws=self.wb["Data"]
        ws['C1']=self.datefin #change la date de mise à jour
        self.wb.save(self.nomfic) #sauvegarde

  

class extraire_forfaits:
    def __init__(self,nom_fichier):
        #charge la feuille des forfaits et utilise la première ligne pour faire ref aux colonnes
        self.feuille=pe.get_sheet(file_name=nom_fichier,name_columns_by_row=0)
        self.liste_vac=self.feuille.column[0] #la colonne 0 contient la liste des vacs
        
    def dic_forfaits(self):
        #renvoie un dictionnaire {vac:[liste des forfaits]}
        dic={}
        for r in self.feuille.rows():
            dic[r[0]]=[r[1],r[2],r[3],r[4]]
        return dic
    

###################################### PROGRAMME ################################################################




plannings=list()
datmaj=date_fin()
forfaits=extraire_forfaits("Forfaits HDC.xlsx")

for p in liste_plannings(datmaj.year): #ouvre chaque planning et l'ajoute à la liste planning
    if "stagiaires" in p:
        plan_stg=stagiaires(p) #s'il s'agit du plannig stagiaires, le charge à part
    else :
        plannings.append(charge_planning(p))

NTrig=nom_trig()
ltrig=NTrig.liste_trig() #liste les trigrammes qui souhaitent utiliser le programme

for t in ltrig:
    majDHC=dhc(t,NTrig.nom_prenom(t),datmaj)
    lvac=list()
    g=NTrig.gamma(t)
    for p in plannings :
        nom=NTrig.nom(t)
        if nom in p.colnames : #si le nom complet est dans les noms de colonnes (planning détachés)
            trig=nom #affecte le nom complet à trig
        else :
            trig=t # sinon laisse le trigramme
        l_vacs=extraire_vac(p,trig,majDHC.date_dhc(),datmaj) #extrait toutes les vacs des plannings
        l_forfaits_trig=crée_liste_hdc(l_vacs,forfaits.dic_forfaits(),g,plan_stg) #crée la liste des forfaits
        lvac.extend(l_forfaits_trig) #crée la liste de liste des dates, vac
    majDHC.export_vers_dhc(lvac)





